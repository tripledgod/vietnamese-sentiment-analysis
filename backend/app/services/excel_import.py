"""
Đọc file Excel nạp danh sách SV/GV (hàng đầu là tiêu đề cột).
"""
import math
import unicodedata
from collections.abc import Set as AbstractSet
from io import BytesIO
from typing import Any

import openpyxl

from app.core.config import DEFAULT_INITIAL_PASSWORD
from app.db import get_db
from app.services.auth_service import hash_password
from app.services.classes_repo import (
    find_class_id_by_name_dept,
    get_class_id_by_class_name,
    get_class_id_numeric,
    ensure_class,
)
from app.services.users import find_by_username, upsert_user

# Tiêu đề cột Excel → map vào users.username (sau _norm_header = dạng ASCII gạch dưới).
_USER_LOGIN_COLUMN_ALIASES: frozenset[str] = frozenset(
    {
        "username",
        "username_mssv",  # sau chuẩn hóa: "username (MSSV)"
        "mssv",
        "ma_sv",
        "ma_so_sinh_vien",
        "ma_so",
        "maso",
        "tai_khoan",
        "login",
    }
)

_ROLE_ALIASES: frozenset[str] = frozenset({"role", "vai_tro", "loai"})

_FULL_NAME_ALIASES: frozenset[str] = frozenset(
    {
        "full_name",
        "ho_ten",
        "ho_va_ten",
        "ten",
        "name",
        "hoten",
    }
)

# Tên lớp / mã lớp (sheet Classes & Users).
_CLASS_COL_ALIASES: frozenset[str] = frozenset(
    {
        "class_name",
        "class_nam",  # lỗi đánh máy phổ biến: thiếu chữ e
        "ten_lop",
        "tenlop",
        "lop",
        "ma_lop",
        "malop",
        "class_code",
    }
)

_DEPARTMENT_ALIASES: frozenset[str] = frozenset(
    {"department", "khoa", "faculty", "don_vi"}
)

_SUBJECT_CODE_ALIASES: frozenset[str] = frozenset(
    {"subject_code", "ma_mon", "mamon", "ma_mh", "subjectcode"}
)

_SUBJECT_NAME_ALIASES: frozenset[str] = frozenset(
    {"subject_name", "ten_mon", "tenmon", "ten_mh", "ten_mon_hoc"}
)


def _norm_header(s: Any) -> str:
    """
    Chuẩn hóa tiêu đề cột Excel: thường + gạch dưới + bỏ dấu tiếng Việt + đ→d.
    Ví dụ: \"Lớp\" → lop, \"Họ và tên\" → ho_va_ten, \"username (MSSV)\" → username_mssv.
    Dấu ngoặc ( ) đổi thành _ để không cắt nhầm đuôi ) của (MSSV).
    """
    if s is None:
        return ""
    t = str(s).strip().lower()
    if not t or t == "none":
        return ""
    for ch in (" ", "-", "\u00a0", "(", ")"):
        t = t.replace(ch, "_")
    while "__" in t:
        t = t.replace("__", "_")
    t = t.strip("_")
    t = "".join(
        c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn"
    )
    return t.replace("đ", "d")


def _parse_role(cell: Any) -> str:
    v = (str(cell).strip().lower() if cell is not None else "") or ""
    if v in ("student", "sinh_vien", "sv", "sinh_viên", "sinh viên"):
        return "student"
    if v in (
        "teacher",
        "lecturer",
        "giang_vien",
        "gv",
        "giảng_viên",
        "giảng viên",
        "giang vien",
    ):
        return "teacher"
    raise ValueError(f"Vai trò không hợp lệ: {cell!r}")


def _col_index(headers: list[str], aliases: AbstractSet[str]) -> int | None:
    for i, h in enumerate(headers):
        if h in aliases:
            return i
    return None


def _cell_int_or_none(val: Any) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return int(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell_login_username(val: Any) -> str:
    """
    MSSV / username từ Excel: ô kiểu Số thành float (vd. 20210001.0) — lưu đúng chuỗi MSSV,
    không phải '20210001.0' (sẽ khiến đăng nhập bằng MSSV thường bị 401).
    """
    if val is None:
        return ""
    if isinstance(val, bool):
        return str(val).strip()
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if math.isfinite(val) and val == int(val):
            return str(int(val))
        return str(val).strip()
    s = str(val).strip()
    if len(s) > 2 and s.endswith(".0"):
        head = s[:-2]
        if head.isdigit():
            return head
    return s


def import_users_from_excel_bytes(
    data: bytes,
    *,
    default_password: str = DEFAULT_INITIAL_PASSWORD,
) -> dict[str, Any]:
    """
    Cột bắt buộc: username / MSSV / mã SV (giá trị = tài khoản đăng nhập) và role (vai_tro).
    Cột khuyến nghị: full_name (ho_ten).
    Lớp: class_id (số) hoặc cặp class_name + department (theo ERD Classes).
    Cột class_code / ma_lop: coi như class_name (department rỗng).
    Cột password / mật khẩu trong file bị bỏ qua — luôn dùng default_password (= DEFAULT_INITIAL_PASSWORD, mặc định Huce@123).
    """
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return {"created": 0, "updated": 0, "errors": ["File trống"]}

        headers = [_norm_header(c) for c in header_row]
        idx_user = _col_index(headers, _USER_LOGIN_COLUMN_ALIASES)
        idx_role = _col_index(headers, _ROLE_ALIASES)
        idx_full = _col_index(headers, _FULL_NAME_ALIASES)
        idx_class_id = _col_index(headers, {"class_id", "id_lop", "lop_id"})
        idx_class = _col_index(headers, _CLASS_COL_ALIASES)
        idx_department = _col_index(headers, _DEPARTMENT_ALIASES)

        errors: list[str] = []
        if idx_user is None:
            errors.append("Thiếu cột MSSV đăng nhập (username, mssv, ma_sv, …)")
        if idx_role is None:
            errors.append("Thiếu cột role / vai_tro")
        if errors:
            return {"created": 0, "updated": 0, "errors": errors}

        created = 0
        updated = 0

        with get_db() as conn:
            for r in rows:
                if r is None or all(v is None or str(v).strip() == "" for v in r):
                    continue
                try:
                    username = _cell_login_username(r[idx_user])
                    if not username:
                        continue
                    role = _parse_role(r[idx_role])
                    full_name = ""
                    if idx_full is not None:
                        full_name = _cell_str(r[idx_full])

                    class_id: int | None = None
                    if idx_class_id is not None:
                        class_id = _cell_int_or_none(r[idx_class_id])
                        if class_id is not None:
                            verified = get_class_id_numeric(conn, class_id)
                            if verified is None:
                                raise ValueError(f"class_id={class_id} không tồn tại trong bảng classes")
                            class_id = verified

                    if class_id is None:
                        cname = ""
                        dept = ""
                        if idx_class is not None:
                            cname = _cell_str(r[idx_class])
                        if idx_department is not None:
                            dept = _cell_str(r[idx_department])
                        if cname:
                            cid = find_class_id_by_name_dept(
                                conn, class_name=cname, department=dept
                            )
                            if cid is None:
                                raise ValueError(
                                    f"Lớp {cname!r} (khoa {dept!r}) chưa có trong bảng classes"
                                )
                            class_id = cid

                    pwd = default_password
                    existed = find_by_username(conn, username) is not None
                    upsert_user(
                        conn,
                        username=username,
                        password_hash=hash_password(pwd),
                        full_name=full_name or username,
                        role=role,
                        class_id=class_id,
                        must_change_password=(role == "student"),
                    )
                    if existed:
                        updated += 1
                    else:
                        created += 1
                except Exception as e:  # noqa: BLE001
                    errors.append(f"Dòng lỗi: {e}")

        return {"created": created, "updated": updated, "errors": errors}
    finally:
        wb.close()


def _classify_master_sheet(headers: list[str]) -> str | None:
    """Nhận diện sheet: subjects | classes | users_master."""
    idx_sub_code = _col_index(headers, _SUBJECT_CODE_ALIASES)
    idx_sub_name = _col_index(headers, _SUBJECT_NAME_ALIASES)
    if idx_sub_code is not None and idx_sub_name is not None:
        return "subjects"

    idx_cn = _col_index(headers, _CLASS_COL_ALIASES)
    idx_dep = _col_index(headers, _DEPARTMENT_ALIASES)
    idx_user_cls = _col_index(headers, _USER_LOGIN_COLUMN_ALIASES)
    # Danh mục lớp: không có cột MSSV (tránh nhầm với sheet Users có thêm department).
    if idx_cn is not None and idx_dep is not None and idx_user_cls is None:
        return "classes"

    idx_user = idx_user_cls
    idx_full = _col_index(headers, _FULL_NAME_ALIASES)
    idx_uclass = _col_index(headers, _CLASS_COL_ALIASES)
    # Sheet Users: có thể thêm cột role / Ghi chú; không bắt buộc. Không trùng fingerprint Subjects.
    if idx_user is not None and idx_uclass is not None and idx_full is not None:
        return "users_master"

    return None


def _assign_master_sheets(wb: Any) -> dict[str, Any]:
    """Map loại sheet -> worksheet (mỗi loại đúng 1 sheet)."""
    found: dict[str, str] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        row0 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row0:
            continue
        headers = [_norm_header(c) for c in row0]
        kind = _classify_master_sheet(headers)
        if not kind:
            continue
        if kind in found:
            raise ValueError(
                f"Có nhiều sheet kiểu {kind!r}: {found[kind]!r} và {sn!r}. "
                "Chỉ để một sheet Subjects, một Classes, một Users."
            )
        found[kind] = sn
    need = {"subjects", "classes", "users_master"}
    missing = need - set(found.keys())
    if missing:
        raise ValueError(
            "File thiếu sheet hoặc tiêu đề cột không khớp. Cần 3 sheet (tiêu đề tiếng Việt có dấu vẫn được, "
            "hệ thống chuẩn hóa tự động):\n"
            "• Subjects: subject_code / mã môn + subject_name / tên môn (Ghi chú tuỳ chọn).\n"
            "• Classes: class_name hoặc Lớp + department / khoa.\n"
            "• Users: MSSV / username / mã SV + Họ và tên (hoặc full_name) + Lớp / class_name.\n"
            f"Thiếu nhận diện: {', '.join(sorted(missing))}."
        )
    return {k: wb[found[k]] for k in found}


def _link_subject_to_all_semesters(conn: Any, subject_id: int) -> int:
    sem_rows = conn.execute("SELECT id FROM semesters").fetchall()
    added = 0
    for r in sem_rows:
        sem_id = int(r["id"])
        before = conn.total_changes
        conn.execute(
            """
            INSERT OR IGNORE INTO semester_subjects (semester_id, subject_id)
            VALUES (?, ?)
            """,
            (sem_id, subject_id),
        )
        if conn.total_changes > before:
            added += 1
    return added


def _upsert_subject_by_code(conn: Any, subject_code: str, subject_name: str) -> str:
    """'created' | 'updated'"""
    code = subject_code.strip()
    name = subject_name.strip()
    if not code:
        raise ValueError("subject_code không được để trống")
    if not name:
        raise ValueError("subject_name không được để trống")
    cur = conn.execute("SELECT id FROM subjects WHERE subject_code = ?", (code,))
    row = cur.fetchone()
    if row:
        sid = int(row["id"])
        conn.execute(
            "UPDATE subjects SET subject_name = ? WHERE id = ?",
            (name, sid),
        )
        return "updated"
    conn.execute(
        "INSERT INTO subjects (subject_name, subject_code) VALUES (?, ?)",
        (name, code),
    )
    return "created"


def import_master_excel_bytes(
    data: bytes,
    *,
    default_password: str = DEFAULT_INITIAL_PASSWORD,
) -> dict[str, Any]:
    """
    Excel 3 sheet (thứ tự tên sheet tuỳ ý, tự nhận theo cột tiêu đề):

    - **Subjects:** subject_code, subject_name, (Ghi chú — bỏ qua).
    - **Classes:** class_name, department.
    - **Users:** cột **username** hoặc **MSSV** (mã đăng nhập cổng SV), full_name, class_name
      (khớp y hệt sheet Classes). Cột password trong Excel bị bỏ qua; mật khẩu lần đầu = default_password
      (DEFAULT_INITIAL_PASSWORD, mặc định Huce@123). Không có cột role — mọi dòng là sinh viên.

    Thứ tự ghi DB: Classes → Subjects (và gắn vào mọi kỳ học đã có) → Users.
    """
    empty_stats = {
        "classes_new": 0,
        "classes_existing": 0,
        "subjects_created": 0,
        "subjects_updated": 0,
        "semester_links_added": 0,
        "users_created": 0,
        "users_updated": 0,
    }
    wb = openpyxl.load_workbook(BytesIO(data), read_only=False, data_only=True)
    try:
        try:
            sheets = _assign_master_sheets(wb)
        except ValueError as e:
            return {"ok": False, **empty_stats, "errors": [str(e)]}
        ws_sub = sheets["subjects"]
        ws_cls = sheets["classes"]
        ws_usr = sheets["users_master"]

        errors: list[str] = []
        stats = dict(empty_stats)

        # --- Sheet Classes ---
        hdr_c = [
            _norm_header(c)
            for c in next(ws_cls.iter_rows(min_row=1, max_row=1, values_only=True), [])
        ]
        rows_c = ws_cls.iter_rows(min_row=2, values_only=True)
        icn = _col_index(hdr_c, _CLASS_COL_ALIASES)
        idep = _col_index(hdr_c, _DEPARTMENT_ALIASES)
        if icn is None or idep is None:
            return {**stats, "ok": False, "errors": ["Sheet Classes thiếu class_name hoặc department"]}

        with get_db() as conn:
            for r in rows_c:
                cn = ""
                if r is None or all(v is None or str(v).strip() == "" for v in r):
                    continue
                try:
                    cn = _cell_str(r[icn])
                    dep = _cell_str(r[idep])
                    if not cn:
                        continue
                    cur = conn.execute(
                        "SELECT 1 FROM classes WHERE class_name = ? AND department = ?",
                        (cn, dep),
                    )
                    existed = cur.fetchone() is not None
                    ensure_class(conn, class_name=cn, department=dep)
                    if existed:
                        stats["classes_existing"] += 1
                    else:
                        stats["classes_new"] += 1
                except Exception as e:  # noqa: BLE001
                    errors.append(f"Classes — dòng lỗi ({cn!r}): {e}")

            # --- Sheet Subjects ---
            hdr_s = [
                _norm_header(c)
                for c in next(ws_sub.iter_rows(min_row=1, max_row=1, values_only=True), [])
            ]
            rows_s = ws_sub.iter_rows(min_row=2, values_only=True)
            isc = _col_index(hdr_s, _SUBJECT_CODE_ALIASES)
            isn = _col_index(hdr_s, _SUBJECT_NAME_ALIASES)
            if isc is None or isn is None:
                return {**stats, "ok": False, "errors": errors + ["Sheet Subjects thiếu subject_code hoặc subject_name"]}

            for r in rows_s:
                code = ""
                if r is None or all(v is None or str(v).strip() == "" for v in r):
                    continue
                try:
                    code = _cell_str(r[isc])
                    sname = _cell_str(r[isn])
                    if not code and not sname:
                        continue
                    kind = _upsert_subject_by_code(conn, code, sname)
                    row = conn.execute(
                        "SELECT id FROM subjects WHERE subject_code = ?",
                        (code.strip(),),
                    ).fetchone()
                    if not row:
                        raise RuntimeError("Không đọc lại được subject sau upsert")
                    sid = int(row["id"])
                    if kind == "created":
                        stats["subjects_created"] += 1
                    else:
                        stats["subjects_updated"] += 1
                    stats["semester_links_added"] += _link_subject_to_all_semesters(conn, sid)
                except Exception as e:  # noqa: BLE001
                    errors.append(f"Subjects — dòng lỗi ({code!r}): {e}")

            # --- Sheet Users (sinh viên) ---
            hdr_u = [
                _norm_header(c)
                for c in next(ws_usr.iter_rows(min_row=1, max_row=1, values_only=True), [])
            ]
            rows_u = ws_usr.iter_rows(min_row=2, values_only=True)
            iu = _col_index(hdr_u, _USER_LOGIN_COLUMN_ALIASES)
            ifu = _col_index(hdr_u, _FULL_NAME_ALIASES)
            iuc = _col_index(hdr_u, _CLASS_COL_ALIASES)
            if iu is None or ifu is None or iuc is None:
                return {
                    **stats,
                    "ok": False,
                    "errors": errors
                    + [
                        "Sheet Users thiếu cột MSSV/đăng nhập (username, mssv, …), full_name hoặc class_name"
                    ],
                }

            for r in rows_u:
                username = ""
                if r is None or all(v is None or str(v).strip() == "" for v in r):
                    continue
                try:
                    username = _cell_login_username(r[iu])
                    if not username:
                        continue
                    full_name = _cell_str(r[ifu])
                    cname = _cell_str(r[iuc])
                    if not cname:
                        raise ValueError("Thiếu class_name")
                    class_id = get_class_id_by_class_name(conn, cname)
                    pwd = default_password
                    existed = find_by_username(conn, username) is not None
                    upsert_user(
                        conn,
                        username=username,
                        password_hash=hash_password(pwd),
                        full_name=full_name or username,
                        role="student",
                        class_id=class_id,
                        must_change_password=True,
                    )
                    if existed:
                        stats["users_updated"] += 1
                    else:
                        stats["users_created"] += 1
                except Exception as e:  # noqa: BLE001
                    errors.append(f"Users — {username!r}: {e}")

        ok = len(errors) == 0
        return {"ok": ok, **stats, "errors": errors}
    finally:
        wb.close()
